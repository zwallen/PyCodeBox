#!/bin/env python
# -*- coding: utf-8 -*-


# ----------------------------------------------------------------#
# Styling Utilities                                               #
# ----------------------------------------------------------------#
# These functions are used for styling data displays and exports. #
# ----------------------------------------------------------------#


def export_styled_xlsx_w_2_headers(df, ws_title, filename):
    """
    Export a pandas DataFrame to a styled Excel file with custom formatting for tables
    with two header rows: a major header row and a sub header row.

    Parameters
    ----------
    df : pandas.DataFrame
        The DataFrame to export. The first two rows should correspond to header and
        sub-header rows.
    ws_title : str
        The title/name for the Excel worksheet.
    filename : str
        The filename (including path if needed) for the output Excel file. Should include
        .xlsx extension.

    Returns
    -------
    None
        Saves a styled Excel workbook to file.

    Features
    --------
    - Custom column widths (55 for first column, then alternating 5/15/20/10 pattern)
    - Left alignment for most cells, left and center alignment for first column
    - First two rows are bold
    - Medium borders on top of first row, bottom of second row, and bottom of last row
    - Thin borders above rows containing 'Total', 'N, %', or 'Mean±SD' in first column
    - Automatic conversion of numeric strings back to numbers where possible
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ws_title

    # Write DataFrame to worksheet (no column names)
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # Set column widths
    col_widths = (
        [(df.iloc[:, 0].str.len().max() + 5)]
        + [5, 15]
        + [5, 15, 20, 10] * (sum(df.iloc[0, :] != '') - 1)
    )
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Apply left alignment to all cells
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Apply center alignment to cells in first column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply bold to first two rows
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(bold=True)

    # Apply medium border to top of row 1, and bottom of row 2 and last row
    for row, border_pos in zip([1, 2, ws.max_row], ['top', 'bottom', 'bottom']):
        for col in range(1, ws.max_column + 1):
            if border_pos == 'top':
                ws.cell(row=row, column=col).border = Border(
                    top=Side(border_style='medium', color='000000')
                )
            else:
                ws.cell(row=row, column=col).border = Border(
                    bottom=Side(border_style='medium', color='000000')
                )

    # Apply thin border below major variable rows and left align
    def grep(pattern):
        import re

        return [
            i + 1
            for i, cell in enumerate(ws['A'])
            if cell.value and re.search(pattern, str(cell.value))
        ]

    border_rows = []
    border_rows += grep('Total')
    border_rows += grep('N, %')
    border_rows += grep('Mean±SD')

    for row in border_rows:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = Border(
                top=Side(border_style='thin', color='000000')
            )
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal='left', vertical='center'
            )

    # Convert numbers from strings back to numbers where possible
    for row in ws.iter_rows(
        min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            if isinstance(cell.value, str):
                try:
                    if '.' in cell.value:
                        cell.value = float(cell.value)
                    else:
                        cell.value = int(cell.value)
                except Exception:
                    pass

    # Save workbook
    wb.save(filename)


def display_styled_table_w_2_headers(df, rows_per_page=50):
    """
    Display a pandas DataFrame as a styled HTML table with custom formatting for tables
    with two header rows: a major header row and a sub header row. Shows data in pages
    with specified number of rows per page, keeping the first two header rows on each
    page.

    Parameters
    ----------
    df : pandas.DataFrame
        The DataFrame to display. The first two rows should correspond to header and
        sub-header rows.
    rows_per_page : int, optional
        Number of data rows to display per page (excluding the 2 header rows)
        (default: 50).

    Returns
    -------
    None
        Displays the styled HTML table.

    Features
    --------
    - First 2 rows are bold and appear on every page
    - Medium border under 2nd row
    - First column cells containing 'Total', '(N, %)', or '(Mean±SD)' are bold
    - No text wrapping
    - Horizontally scrollable if needed
    - Navigation buttons for page switching
    - Quarto-compatible styling
    """
    from IPython.display import display, HTML
    import math

    # Extract header rows (first 2 rows) and data rows
    header_rows = df.iloc[:2]
    data_rows = df.iloc[2:]

    # Calculate total number of pages
    total_pages = math.ceil(len(data_rows) / rows_per_page) if len(data_rows) > 0 else 1

    # Generate unique ID for this table instance
    import uuid

    table_id = str(uuid.uuid4())[:8]

    # Create JavaScript for pagination and cell styling
    js_script = f"""
    <script>
    document.addEventListener('DOMContentLoaded', function() {{
      let currentPage_{table_id} = 1;
      const totalPages_{table_id} = {total_pages};
      const rowsPerPage_{table_id} = {rows_per_page};
    
      // Function to bold specific cells in column 1
      function styleFirstColumn_{table_id}() {{
        const rows = document.querySelectorAll('#table_{table_id} tr');
        rows.forEach(row => {{
          const firstCell = row.querySelector('td:first-child');
          if (firstCell) {{
            const cellText = firstCell.textContent || firstCell.innerText;
            if (cellText.includes('Total') || cellText.includes('(N, %)') || cellText.includes('(Mean±SD)')) {{
              firstCell.style.fontWeight = 'bold';
            }}
          }}
        }});
      }}
    
      // Function to show specific page
      function showPage_{table_id}(pageNum) {{
        // Hide all data rows
        const dataRows = document.querySelectorAll('#table_{table_id} .data-row');
        dataRows.forEach(row => row.style.display = 'none');
      
        // Show rows for current page
        const startIdx = (pageNum - 1) * rowsPerPage_{table_id};
        const endIdx = startIdx + rowsPerPage_{table_id};
      
        for (let i = startIdx; i < endIdx && i < dataRows.length; i++) {{
          dataRows[i].style.display = 'table-row';
        }}
      
        // Update page info
        document.getElementById('page-info_{table_id}').textContent = 
          `Page ${{pageNum}} of ${{totalPages_{table_id}}}`;
      
        // Update button states
        document.getElementById('prev-btn_{table_id}').disabled = (pageNum === 1);
        document.getElementById('next-btn_{table_id}').disabled = (pageNum === totalPages_{table_id});
      
        currentPage_{table_id} = pageNum;
      
        // Style first column after showing page
        styleFirstColumn_{table_id}();
      }}
    
      // Event listeners for pagination buttons
      document.getElementById('prev-btn_{table_id}').addEventListener('click', function() {{
        if (currentPage_{table_id} > 1) {{
          showPage_{table_id}(currentPage_{table_id} - 1);
        }}
      }});
    
      document.getElementById('next-btn_{table_id}').addEventListener('click', function() {{
        if (currentPage_{table_id} < totalPages_{table_id}) {{
          showPage_{table_id}(currentPage_{table_id} + 1);
        }}
      }});
    
      // Show first page initially
      showPage_{table_id}(1);
    }});
    </script>
    """

    # Create HTML table structure
    html_parts = []

    # Add header rows (these will always be visible)
    for idx, (_, row) in enumerate(header_rows.iterrows()):
        row_class = 'header-row'
        cells = ''.join([f'<td>{cell}</td>' for cell in row.values])
        html_parts.append(f'<tr class="{row_class}">{cells}</tr>')

    # Add data rows (these will be shown/hidden based on pagination)
    for idx, (_, row) in enumerate(data_rows.iterrows()):
        row_class = 'data-row'
        cells = ''.join([f'<td>{cell}</td>' for cell in row.values])
        html_parts.append(f'<tr class="{row_class}">{cells}</tr>')

    # Combine all rows
    table_rows = '\n'.join(html_parts)

    # Create pagination controls
    pagination_controls = f"""
    <div class="pagination-controls" style="margin: 10px 0; text-align: center;">
      <button id="prev-btn_{table_id}" style="margin: 0 10px; padding: 5px 10px;">← Previous</button>
      <span id="page-info_{table_id}" style="margin: 0 10px; font-weight: bold;">Page 1 of {total_pages}</span>
      <button id="next-btn_{table_id}" style="margin: 0 10px; padding: 5px 10px;">Next →</button>
    </div>
    """

    # Full CSS: Remove wrapping, ensure column expansion, remove width restrictions
    styled_html = f"""
    {js_script}
    <style>
    /* Remove Quarto's max-width */
    .main-container {{
      max-width: none !important;
    }}
    /* Style the table */
    #table_{table_id} {{
      width: 100%;
      border-collapse: collapse;
      table-layout: auto;
    }}
    /* Style each cell to prevent wrapping */
    #table_{table_id} td {{
      white-space: nowrap;
      padding: 8px 12px;
      font-size: 14px;
      border: 1px solid #ddd;
    }}
    /* Make header rows bold */
    #table_{table_id} .header-row td {{
      font-weight: bold;
    }}
    /* Add medium border under 2nd header row */
    #table_{table_id} .header-row:nth-child(2) td {{
      border-bottom: 2px solid #333;
    }}
    /* Make table horizontally scrollable if necessary */
    .table-wrapper {{
      overflow-x: auto;
    }}
    /* Style pagination controls */
    .pagination-controls button {{
      background-color: #f8f9fa;
      border: 1px solid #dee2e6;
      border-radius: 4px;
      cursor: pointer;
    }}
    .pagination-controls button:hover:not(:disabled) {{
      background-color: #e9ecef;
    }}
    .pagination-controls button:disabled {{
      background-color: #e9ecef;
      color: #6c757d;
      cursor: not-allowed;
    }}
    </style>
    <div class='table-wrapper'>
      <table id="table_{table_id}">
        {table_rows}
      </table>
    </div>
    {pagination_controls}
    """

    # Display the styled HTML table
    display(HTML(styled_html))
